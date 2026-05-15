# -*- coding: utf-8 -*-
"""Cập nhật sheet 02 + 05: Loại buổi A/B/C; bỏ band-centric trên đánh giá tháng."""
from pathlib import Path

import openpyxl

BASE = Path(__file__).resolve().parent.parent


def _tracking_xlsx(index: int) -> Path:
    xs = sorted((BASE / "02 - Theo Dõi Tiến Độ").glob("*.xlsx"))
    return xs[index]


def patch_buoi():
    f = _tracking_xlsx(0)
    wb = openpyxl.load_workbook(f)
    ws = wb.active
    if ws["B5"].value != "Loại buổi (A/B/C)":
        ws.insert_cols(2)
        ws["B5"] = "Loại buổi (A/B/C)"
    ws["B2"] = (
        "Phụ thuộc: Giáo Án — ghi A (ngữ pháp+drill), B (nghe+nói), C (đọc+viết+ôn từ)."
    )
    wb.save(f)
    wb.close()
    print("patched", f.name)


def patch_tuan():
    f = _tracking_xlsx(1)
    wb = openpyxl.load_workbook(f)
    ws = wb.active
    h8 = ws.cell(5, 8).value
    if h8 and "Kỹ năng" in str(h8):
        ws.cell(5, 8, "Trọng tâm tuần (A/B/C + kỹ năng)")
    if ws.cell(5, 14).value != "IELTS / band (chỉ Năm 2 — tùy chọn)":
        ws.insert_cols(14)
        ws.cell(5, 14, "IELTS / band (chỉ Năm 2 — tùy chọn)")
    wb.save(f)
    wb.close()
    print("patched", f.name)


def patch_thang():
    f = _tracking_xlsx(2)
    wb = openpyxl.load_workbook(f)
    ws = wb.active
    ws["G5"] = "Nghe — mức nền (ghi chú / 1-5)"
    ws["H5"] = "Đọc — mức nền (ghi chú / 1-5)"
    ws["I5"] = "Viết — mức nền (ghi chú / 1-5)"
    ws["J5"] = "Nói — mức nền (ghi chú / 1-5)"
    ws["K5"] = "Tổng hợp nền (band IELTS: chỉ Năm 2 nếu cần)"
    ws["O5"] = "Điều chỉnh kế hoạch nền / Năm 2"
    if ws["P5"].value:
        ws["P5"] = "Mục tiêu năng lực tháng sau"
    ws["A3"] = (
        "Cung cấp cho: Báo cáo tháng gửi học viên | Điều chỉnh kế hoạch. "
        "Năm 1: rubric nền — không ép band hàng tháng."
    )
    wb.save(f)
    wb.close()
    print("patched", f.name)


def patch_vocab():
    d = BASE / "05 - Từ Vựng & Ngữ Pháp"
    f = next(d.glob("*.xlsx"))
    wb = openpyxl.load_workbook(f)
    ws = wb[wb.sheetnames[0]]
    if ws["L6"].value != "Loại buổi (A/B/C)":
        ws.insert_cols(12)
        ws["L6"] = "Loại buổi (A/B/C)"
    wb.save(f)
    wb.close()
    print("patched", f.name)


def main():
    patch_buoi()
    patch_tuan()
    patch_thang()
    patch_vocab()


if __name__ == "__main__":
    main()
